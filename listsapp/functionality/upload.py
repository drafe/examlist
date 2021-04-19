import logging

import attr
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as indx

from listsapp.models import Subject

log_xslx = logging.getLogger(__name__)


@attr.s
class Parser:
    """" Класс, реализующий получение данных из xslx файла
    конструктор:
    file : str - имя файла для парсинга
    wsh : str - имя листа в файле, на которм находятся данные
    rule : json - правила, по которому нужно парсить
    """
    # RULES = {'columns': {'cipher': 'A', 'subjects': 'B', 'departments': 'BG',
    #                                'controls': {'exam': 'C', 'quiz': 'D'},
    #                                '1_sem': 'R', 'lectures': 1, 'practice': 2, 'labs': 3},
    #           'ciphers': ('ОНБ', 'ПБ')
    #          }
    # todo: хранить не буквы, а колонки?
    # todo отловка пустого названия листа
    _file = attr.ib(type=str)
    _wsh = attr.ib(type=str)
    _rule = attr.ib(type=dict, repr=False)
    _data = attr.ib(type=list, init=False, repr=False)

    def parse(self):
        """ Парсит xslx-файл по правилу json
            на список предметов,
            список кафедр, которые читают предметы,
            списки строк учебного плана, в которых название предмета заменено на его индекс из списка предметов
            ___________
            словарь вида
            {"subjects": subjects, - список предметов
            "exam": exams, - список экзаменнационных предметов
            (индекс предмета, семестр, часы лекций, лабораторок, практики)
            "quiz": quiz, - список зачетных предметов (аналогично)
            "m_qu": main_quiz} - список предметов с диф.зачётом
            """

        log_xslx.debug('start')
        self.__load_data()
        log_xslx.debug('ending')
        return self.__clear_data()

    def __load_data(self):
        """ Выгружает данные из xlsx-файла в кучу """
        log_xslx.debug('start')
        wb = load_workbook(filename=self._file, read_only=True, data_only=True)
        sh = wb[self._wsh]
        rows = sh.max_row
        cols = self._rule['columns']['departments']
        slc = sh['A1':f"{cols}{str(rows)}"]
        self._data = [[str(_.value) for _ in s] for s in slc]
        wb.close()
        del slc, sh, wb
        log_xslx.debug('end')

    @staticmethod
    def __is_subject(row: list, ciph, ciphers, subj, sem_exam, sem_quiz):
        return row[ciph].split('.', 1)[0] in ciphers \
               and ((row[subj] not in ('0', '', 'None', None)) \
               and ((row[sem_exam] not in ('0', '', 'None', None)) \
               or (row[sem_quiz] not in ('0', '', 'None', None))))

    def __clear_data(self):
        """ Очищает данные, готовит словарь вида
        {"subjects": subjects, - список предметов
        "exam": exams, - список экзаменнационных предметов
        (индекс предмета, семестр, часы лекций, лабораторок, практики)
        "quiz": quiz, - список зачетных предметов (аналогично)
        "m_qu": main_quiz} - список предметов с диф.зачётом
        """
        import re

        log_xslx.debug('start')
        ciphers = self._rule['ciphers']
        cols = self._rule['columns']
        ciph, subj, deps = indx(cols['cipher']) - 1, indx(cols['subjects']) - 1, indx(cols['departments']) - 1
        lec, lab, pract = cols['lectures'], cols['labs'], cols['practice']
        f_sem = indx(cols['1_sem']) - 1
        sem_exam = indx(cols['controls']['exam']) - 1
        sem_quiz = indx(cols['controls']['quiz']) - 1

        data = list(filter(lambda x: self.__is_subject(x, ciph, ciphers, subj, sem_exam, sem_quiz),
                           self._data))
        self._data = data
        subjects = [_[subj] for _ in data]

        l_d = len(data)
        exams = [[i, s, data[i][f_sem + (int(s) - 1) * 4 + lec], data[i][f_sem + (int(s) - 1) * 4 + lab],
                  data[i][f_sem + (int(s) - 1) * 4 + pract]]
                 for i in range(l_d) for s in re.split('[,.]', data[i][sem_exam]) if s != 'None']
        exams = [list(map(lambda x: int(x) if x != 'None' else 0, _)) for _ in exams]
        quiz = [[i, s, data[i][f_sem + (int(s) - 1) * 4 + lec], data[i][f_sem + (int(s) - 1) * 4 + lab],
                 data[i][f_sem + (int(s) - 1) * 4 + pract]]
                for i in range(l_d) for s in re.split('[,.]', data[i][sem_quiz]) if s != 'None' and s[-1] != '*']
        quiz = [list(map(lambda x: int(x) if x != 'None' else 0, _)) for _ in quiz]
        main_quiz = [[i, s[0], data[i][f_sem + (int(s[0]) - 1) * 4 + lec], data[i][f_sem + (int(s[0]) - 1) * 4 + lab],
                      data[i][f_sem + (int(s[0]) - 1) * 4 + pract]]
                     for i in range(l_d) for s in re.split('[,.]', data[i][sem_quiz]) if s != 'None' and s[-1] == '*']
        main_quiz = [list(map(lambda x: int(x) if x != 'None' else 0, _)) for _ in main_quiz]

        log_xslx.debug('end')

        return {"subjects": subjects,
                "exam": exams, "quiz": quiz, "m_qu": main_quiz}


class FuzzySubjectsComparison:
    """
    Класс, реализующий сравнение названий дисциплин для новой специальности с существующими дисциплинами из базы
    ____________
    конструктор:
    compare_sub : list[str] - список названий дисциплин
    ____________
    формат выхода:
    similar_sub : dict[ subject: similar_subjects ] - словарь с ключами-названиями дисциплин
    и значениями-листами пар (оценка; объект Subject)
    """
    MIN_WORLD_LENGTH = 3
    THRESHOLD_SENTENCE = 0.25
    THRESHOLD_WORD = 0.45
    NGRAM_LENGTH = 2

    def __init__(self, compare_sub: list):
        self.compare = compare_sub
        self.current = Subject.objects.all()

    def compareAll(self):
        d = [sorted(list(filter(lambda x: x[0] > 0, [(self.getFuzzyEqualValue(i, j.subject), j) for j in self.current])),
                    key=lambda x: x[0], reverse=True)
             for i in self.compare]
        return d

    @staticmethod
    def normal(s: str):
        return "".join(c for c in s if c.isalnum()).lower()

    @staticmethod
    def words(s: str):
        f = FuzzySubjectsComparison
        return [f.normal(_) for _ in s.split(' ') if len(_) >= f.MIN_WORLD_LENGTH]

    @staticmethod
    def isWordsFuzzyEqual(first: str, second: str):
        if first == second:
            return True

        f = FuzzySubjectsComparison
        equalNgram = 0
        first_gram_count = len(first) - f.NGRAM_LENGTH + 1
        second_gram_count = len(second) - f.NGRAM_LENGTH + 1
        usedNgram = [False for _ in range(second_gram_count)]

        for i in range(first_gram_count):
            first_gram = first[i:i + f.NGRAM_LENGTH]
            for j in range(second_gram_count):
                if not usedNgram[j]:
                    second_gram = second[j:j + f.NGRAM_LENGTH]
                    if first_gram == second_gram:
                        equalNgram += 1
                        usedNgram[j] = True
                        break
        tanimoto = equalNgram / (first_gram_count + second_gram_count - equalNgram)
        return tanimoto >= f.THRESHOLD_WORD

    @staticmethod
    def getFuzzyEqualValue(first: str, second: str):
        if not first and not second:
            return 1.0

        if not first or not second:
            return 0.0

        f = FuzzySubjectsComparison
        equalWords = 0
        first_words = f.words(first)
        second_words = f.words(second)
        usedWords = [False for _ in range(len(second_words))]

        for i in first_words:
            for _, j in enumerate(second_words):
                if not usedWords[_]:
                    if f.isWordsFuzzyEqual(i, j):
                        equalWords += 1
                        usedWords[_] = True
                        break

        return equalWords / (len(first_words) + len(second_words) - equalWords)


if '__name__' == '__main__':
    s = ['Иностранный язык', 'История', 'Философия', 'Иностранный язык в профессиональной сфере',
         'Математика', 'Физика', 'Информатика и информационно-коммуникационные технологии',
         'Правоведение', 'Экономика', 'Дискретная математика',
         'Математическая логика',
         'История науки и техники / Культурология', 'Методика преподавания / Педагогика',
         'Социология и политология / Религиоведение', 'Операционные системы',
         'Электротехника, электроника и схемотехника', 'Базы данных', 'Сети и телекоммуникации',
         'Основы программирования', 'Безопасность жизнедеятельности', 'Основы охраны труда',
         'Программирование', 'Web-программирование',
         'СУБД Oracle', 'Объектно-ориентированное программирование',
         'Современные информационные системы и технологии', 'Инженерная и компьютерная графика',
         'Архитектура ЭВМ и микроконтроллеров', 'ЭВМ и периферийные устройства',
         'Программирование в системе "1С: Предприятие" / Администрирование системы "1С: Предприятие" / Компьютерный дизайн',
         'Вычислительная математика  / Численные методы /Вычислительные методы',
         'Программирование робототехнических систем / Администрирование операционных систем /  Программные средства обработки графической информации',
         'Интернет-технологии /  Аппаратные средства локальных сетей / Web-дизайн',
         'Программирование в Unix / Администрирование распределённых систем / Компьютерная анимация и видео']
    print(FuzzySubjectsComparison(s).compareAll())
