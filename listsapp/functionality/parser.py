import logging

import attr
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string as indx

from listsapp.models import Subject

log_xslx = logging.getLogger(__name__)


@attr.s
class Parser:
    """" Класс, реализующий получение данных из xslx файла
    конструктор:
    file : str - имя файла для парсинга
    wsh : str - имя листа в файле, на которм находятся данные
    rule : json - правила, по которому нужно парсить
    """
    # RULES = {'columns': {'cipher': 'A', 'subjects': 'B', 'departments': 'BG',
    #                                'controls': {'exam': 'C', 'quiz': 'D'},
    #                                '1_sem': 'R', 'lectures': 1, 'practice': 2, 'labs': 3},
    #           'ciphers': ('ОНБ', 'ПБ')
    #          }
    # todo: хранить не буквы, а колонки?
    # todo отловка пустого названия листа
    _file = attr.ib(type=str)
    _wsh = attr.ib(type=str)
    _rule = attr.ib(type=dict, repr=False)
    _data = attr.ib(type=list, init=False, repr=False)

    def parse(self):
        """ Парсит xslx-файл по правилу json
            на список предметов,
            список кафедр, которые читают предметы,
            списки строк учебного плана, в которых название предмета заменено на его индекс из списка предметов
            ___________
            словарь вида
            {"subjects": subjects, - список предметов
            "exam": exams, - список экзаменнационных предметов
            (индекс предмета, семестр, часы лекций, лабораторок, практики)
            "quiz": quiz, - список зачетных предметов (аналогично)
            "m_qu": main_quiz} - список предметов с диф.зачётом
            """

        log_xslx.debug('start')
        self.__load_data()
        log_xslx.debug('ending')
        return self.__clear_data()

    def __load_data(self):
        """ Выгружает данные из xlsx-файла в кучу """
        log_xslx.debug('start')
        wb = load_workbook(filename=self._file, read_only=True, data_only=True)
        sh = wb[self._wsh]
        rows = sh.max_row
        cols = self._rule['columns']['departments']
        slc = sh['A1':f"{cols}{str(rows)}"]
        self._data = [[str(_.value) for _ in s] for s in slc]
        wb.close()
        del slc, sh, wb
        log_xslx.debug('end')

    @staticmethod
    def __is_subject(row: list, ciph, ciphers, subj, sem_exam, sem_quiz):
        return row[ciph].split('.', 1)[0] in ciphers \
               and ((row[subj] not in ('0', '', 'None', None)) \
               and ((row[sem_exam] not in ('0', '', 'None', None)) \
               or (row[sem_quiz] not in ('0', '', 'None', None))))

    def __clear_data(self):
        """ Очищает данные, готовит словарь вида
        {"subjects": subjects, - список предметов
        "exam": exams, - список экзаменнационных предметов
        (индекс предмета, семестр, часы лекций, лабораторок, практики)
        "quiz": quiz, - список зачетных предметов (аналогично)
        "m_qu": main_quiz} - список предметов с диф.зачётом
        """
        import re

        log_xslx.debug('start')
        ciphers = self._rule['ciphers']
        cols = self._rule['columns']
        ciph, subj, deps = indx(cols['cipher']) - 1, indx(cols['subjects']) - 1, indx(cols['departments']) - 1
        lec, lab, pract = cols['lectures'], cols['labs'], cols['practice']
        f_sem = indx(cols['1_sem']) - 1
        sem_exam = indx(cols['controls']['exam']) - 1
        sem_quiz = indx(cols['controls']['quiz']) - 1

        data = list(filter(lambda x: self.__is_subject(x, ciph, ciphers, subj, sem_exam, sem_quiz),
                           self._data))
        self._data = data
        subjects = [_[subj] for _ in data]

        l_d = len(data)
        exams = [[i, s, data[i][f_sem + (int(s) - 1) * 4 + lec], data[i][f_sem + (int(s) - 1) * 4 + lab],
                  data[i][f_sem + (int(s) - 1) * 4 + pract]]
                 for i in range(l_d) for s in re.split('[,.]', data[i][sem_exam]) if s != 'None']
        exams = [list(map(lambda x: int(x) if x != 'None' else 0, _)) for _ in exams]
        quiz = [[i, s, data[i][f_sem + (int(s) - 1) * 4 + lec], data[i][f_sem + (int(s) - 1) * 4 + lab],
                 data[i][f_sem + (int(s) - 1) * 4 + pract]]
                for i in range(l_d) for s in re.split('[,.]', data[i][sem_quiz]) if s != 'None' and s[-1] != '*']
        quiz = [list(map(lambda x: int(x) if x != 'None' else 0, _)) for _ in quiz]
        main_quiz = [[i, s[0], data[i][f_sem + (int(s[0]) - 1) * 4 + lec], data[i][f_sem + (int(s[0]) - 1) * 4 + lab],
                      data[i][f_sem + (int(s[0]) - 1) * 4 + pract]]
                     for i in range(l_d) for s in re.split('[,.]', data[i][sem_quiz]) if s != 'None' and s[-1] == '*']
        main_quiz = [list(map(lambda x: int(x) if x != 'None' else 0, _)) for _ in main_quiz]

        log_xslx.debug('end')

        return {"subjects": subjects,
                "exam": exams, "quiz": quiz, "m_qu": main_quiz}


